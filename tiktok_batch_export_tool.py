# ============ 只留最基础启动必须 ============
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
import re
import threading
import json
import os
import sys
import time
from datetime import datetime, timedelta

# 全局缓存
url_cache = {}
client = None  # 延迟创建

# ========================== 图标设置 ==========================
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

APP_ICON = resource_path("logo.ico")

# ====================== 授权相关 ======================
AUTH_FILE = os.path.join(os.path.expanduser("~"), ".tiktok_ads_auth.dat")
RULE_FILE = os.path.join(os.path.expanduser("~"), ".tiktok_ads_rules.json")
QQ_NUM = "349163112"
WECHAT_ID = "CloudPark3000"
APP_TITLE = "TikTok投流解析导出飞书工具 v1.1"

TIKTOK_SHORT_PATTERN = re.compile(r"https://www\.tiktok\.com/t/\S+")
ADS_CODE_PATTERN = re.compile(r"#[\w/=+]+")
VID_PATTERN = re.compile(r'/video/(\d+)')

class AuthManager:
    def __init__(self):
        self.valid = False
        self.expire_time = None
        self.auth_type = None
        self.check_auth()

    def check_auth(self):
        if not os.path.exists(AUTH_FILE):
            self.valid = False
            return
        try:
            with open(AUTH_FILE, "r", encoding="utf-8") as f:
                auth_data = json.load(f)
            self.auth_type = auth_data.get("type")
            expire_str = auth_data.get("expire")
            if self.auth_type == "forever":
                self.valid = True
                return
            self.expire_time = datetime.strptime(expire_str, "%Y-%m-%d %H:%M:%S")
            self.valid = datetime.now() < self.expire_time
        except:
            self.valid = False

    def save_auth(self, auth_type, expire_dt=None):
        data = {"type": auth_type}
        if expire_dt:
            data["expire"] = expire_dt.strftime("%Y-%m-%d %H:%M:%S")
        with open(AUTH_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f)
        self.check_auth()

    def get_remain_days(self):
        if self.auth_type == "forever":
            return "永久有效"
        if not self.expire_time:
            return "未激活"
        return f"剩余 {max(0, (self.expire_time - datetime.now()).days)} 天"


class ActivateWindow(tk.Toplevel):
    def __init__(self, parent, auth_manager):
        super().__init__(parent)
        self.parent = parent
        self.auth_mgr = auth_manager
        self.withdraw()
        self.title("软件激活")
        self.geometry("520x260")
        self.resizable(0, 0)
        self.transient(parent)
        self.grab_set()

        try:
            self.iconbitmap(APP_ICON)
        except:
            pass

        ttk.Label(self, text="软件未激活或已过期", font=("微软雅黑", 12)).pack(pady=10)
        ttk.Label(self, text=f"售后：QQ {QQ_NUM} | 微信 {WECHAT_ID}", foreground="red").pack()

        ttk.Label(self, text="激活码：").place(x=30, y=90)
        self.code_input = ttk.Entry(self, width=55)
        self.code_input.place(x=90, y=90)

        ttk.Button(self, text="立即激活", command=self.do_activate).place(x=180, y=140, width=140)
        ttk.Button(self, text="退出", command=self.quit_app).place(x=180, y=180, width=140)
        self.after(10, self.deiconify)

    def do_activate(self):
        code = self.code_input.get().strip()
        if len(code) < 12:
            messagebox.showerror("错误", "激活码格式不正确")
            return
        now = datetime.now()
        if code.startswith("DAY1_"):
            self.auth_mgr.save_auth("day1", now + timedelta(days=1))
        elif code.startswith("DAY3_"):
            self.auth_mgr.save_auth("day3", now + timedelta(days=3))
        elif code.startswith("WEEK_"):
            self.auth_mgr.save_auth("week", now + timedelta(days=7))
        elif code.startswith("MONTH_"):
            self.auth_mgr.save_auth("month", now + timedelta(days=30))
        elif code.startswith("QUARTER_"):
            self.auth_mgr.save_auth("quarter", now + timedelta(days=90))
        elif code.startswith("YEAR_"):
            self.auth_mgr.save_auth("year", now + timedelta(days=365))
        elif code.startswith("FOREVER_"):
            self.auth_mgr.save_auth("forever")
        else:
            messagebox.showerror("失败", "激活码无效")
            return
        messagebox.showinfo("成功", "激活成功！请重启软件")
        self.destroy()

    def quit_app(self):
        self.destroy()
        self.parent.destroy()


# ====================== 规则管理相关 ======================
class RuleManager:
    def __init__(self):
        self.rules = {}
        self.global_product_map = {}  # 1. 先初始化空字典
        self.load_rules()  # 2. 再加载规则
        self.init_default_rules()  # 3. 再初始化默认规则
        self.build_global_map()  # 4. 最后再构建全局映射

    def load_rules(self):
        if os.path.exists(RULE_FILE):
            try:
                with open(RULE_FILE, "r", encoding="utf-8") as f:
                    self.rules = json.load(f)
            except:
                self.rules = {}

    def save_rules(self):
        with open(RULE_FILE, "w", encoding="utf-8") as f:
            json.dump(self.rules, f, ensure_ascii=False, indent=2)
        # 保存后重建全局映射
        self.build_global_map()

    def build_global_map(self):
        """构建全局产品映射：遍历所有商家规则，用于全局匹配产品"""
        self.global_product_map.clear()
        for merchant, rule in self.rules.items():
            prod_map = rule.get("product_map", {})
            for raw_name, new_name in prod_map.items():
                # 后定义的同名称会覆盖，保证唯一性
                self.global_product_map[raw_name] = (merchant, new_name)

    def init_default_rules(self):
        default_rules = {
            "ycz": {
                "product_map": {
                    "hy": "YCZ海洋", "jh": "YCZ借火", "jlb": "YCZ俱乐部", "kh": "YCZ狂欢",
                    "wm": "乌木", "qf": "清风", "hp": "琥珀", "tk": "天空",
                    "hyhp": "海洋+琥珀", "hphy": "海洋+琥珀", "hywm": "海洋+乌木", "wmhy": "海洋+乌木",
                    "khjlb": "狂欢+俱乐部", "jlbkh": "狂欢+俱乐部", "hykh": "海洋+狂欢", "khhy": "海洋+狂欢",
                    "hyjh": "海洋+借火", "jhhy": "海洋+借火", "hyjlb": "海洋+俱乐部", "jlbhy": "海洋+俱乐部",
                    "khjh": "狂欢+借火", "jhkh": "狂欢+借火", "hytk": "海洋+天空", "tkhy": "海洋+天空",
                    "qfw mhp": "清风+乌木+琥珀", "hyjlbkh": "海洋+俱乐部+狂欢",
                    "hyjlbkjh": "海洋+俱乐部+借火", "ycznv": "YCZ女士香水套盒",
                    "ycznan": "YCZ男士香水套盒", "hyjlbjhkh": "海洋+俱乐部+借火+狂欢",
                    "hytz": "海洋+清风+乌木+琥珀"
                },
                "feishu_url": "https://rcnj71ddyxpv.feishu.cn/wiki/JygdwS0KriMWOikh69ecRCdEnrf",
                "headers": ["日期", "ID", "产品名称", "链接", "VID", "code", "挂链接店铺名称"]
            },
            "nto": {
                "product_map": {"yuan": "1730393768499712804", "fang": "1729452323760608036"},
                "feishu_url": "https://rcnj71ddyxpv.feishu.cn/wiki/LPjKwL7jHixtIxkRFIlzcwAien0e",
                "headers": ["日期", "handle", "PID(or型号)", "链接", "vid", "code"]
            },
            "pritom": {
                "product_map": {"pritom": ""},
                "feishu_url": "https://rcnj71ddyxpv.feishu.cn/wiki/H75awTGNOiHVmRktFr4ckShEnz4",
                "headers": ["日期", "handle", "链接", "vid", "code"]
            },
            "zdeer": {
                "product_map": {"zdeer.l": "口喷金", "zdeer.j": "口喷绿", "zdeer.l.j": ""},
                "feishu_url": "https://rcnj71ddyxpv.feishu.cn/wiki/HOQCwn15IimdHiktp76cIXEHnVf",
                "headers": ["账号ID", "产品名称", "发布日期", "视频链接", "投流码"]
            },
            "shokz": {
                "product_map": {
                    "bos": "BOS", "Bos": "BOS", "thk": "THK", "Thk": "THK",
                    "wharf": "Wharf", "Wharf": "Wharf", "cct": "CCT", "Cct": "CCT",
                    "hc": "HC", "Hc": "HC", "se": "SE", "Se": "SE",
                    "seal": "SEAL", "Seal": "SEAL", "usc+": "UCS+", "Usc+": "UCS+",
                    "usc": "UCS", "Usc": "UCS", "cws": "CWS", "Cws": "CWS",
                    "vie": "VIE", "Vie": "VIE", "nt": "NT", "Nt": "NT",
                    "being": "BEING", "Being": "BEING", "lok": "LOK", "Lok": "LOK"
                },
                "feishu_url": "https://rcni71ddyxpv.feishu.cn/wiki/QWZCwfoEQiUOAOkufsecD8TFnNb",
                "headers": ["日期", "ID", "产品名称", "链接", "VID"]
            },
            "xwmus": {
                "product_map": {"xwmus": "", "xwm": "", "Xwm": ""},
                "feishu_url": "https://rcnj71ddyxpv.feishu.cn/wiki/GGvrwriugiRijqkMoKfc5kxVnkh",
                "headers": ["日期", "handle", "链接", "vid", "code"]
            },
            "drizzly": {
                "product_map": {"drizzly": "胶囊充电宝", "jn": "胶囊充电宝", "driz": "胶囊充电宝"},
                "feishu_url": "https://rcnj71ddyxpv.feishu.cn/wiki/PTWDw20nbiIU4Uk20SHchWYjnJd",
                "headers": ["日期", "ID", "产品名称", "链接", "VID", "code"]
            },
            "skg": {
                "product_map": {
                    "Gs500n": "GS500-N", "gs500n": "GS500-N", "Gs500-n": "GS500-N", "GS500-N": "GS500-N",
                    "k5-3": "K5-3", "K5-3": "K5-3", "g7promax": "G7 PROMAX", "G7promax": "G7 PROMAX",
                    "G7 promax": "G7 PROMAX", "G7 Promax": "G7 PROMAX", "G7Proflod": "G7 PRO-FOLD", "G7proflod": "G7 PRO-FOLD"
                },
                "feishu_url": "https://rcnj71ddyxpv.feishu.cn/wiki/Mz39wh07qi97dgkfq6mc3Ispnjb",
                "headers": ["日期", "ID", "产品名称", "链接", "VID", "code"]
            },
            "Sheet9": {
                "product_map": {
                    "q30": "A3028", "Q30": "A3028", "a3028": "A3028", "A3028": "A3028",
                    "p30i": "A3959", "P30i": "A3959", "a3959": "A3959", "A3959": "A3959",
                    "A31X1": "A31X1", "a31x1": "A31X1"
                },
                "feishu_url": "https://rcnj71ddyxpv.feishu.cn/wiki/RRGSwTbLxiOH24kOLZIcScf7nRg",
                "headers": ["日期", "ID", "产品名称", "链接", "VID"]
            },
            "Sheet10": {
                "product_map": {
                    "9215": "A9215", "1638": "A1638", "1614": "A1614", "1665": "A1665", "1695": "A1695",
                    "121d": "121D", "121D": "121D", "91C8": "91C8", "91c8": "91C8",
                    "1695 121D": "A1695+121D", "121d 1695": "A1695+121D", "121D 1695": "A1695+121D",
                    "1695 121d": "A1695+121D", "a2697": "A2697", "A2697": "A2697",
                    "1614 1695": "1614+1695", "1695 1614": "1614+1695", "9196": "9196",
                    "1614 9215": "1614+9215", "9215 1614": "1614+9215",
                    "121d 9196": "121D+9196", "121D 9196": "121D+9196",
                    "9196 121d": "121D+9196", "9196 121D": "121D+9196",
                    "2697 1638 1695": "2697+1638+1695", "1638 2697 1638": "2697+1638+1695",
                    "1638 1695 2697": "2697+1638+1695", "1695 1638 2697": "2697+1638+1695",
                    "1695 1695 1695": "A1695", "1695 9196 1618": "A1695+121D+9196",
                    "9196 121d 1695": "A1695+121D+9196", "9196 121D 1695": "A1695+121D+9196",
                    "121d 9196 1695": "A1695+121D+9196", "121D 9196 1695": "A1695+121D+9196",
                    "121d 1614 9196": "121D+91C8", "91c8 121d": "91C8+121D",
                    "91C8 121D": "91C8+121D", "121d 91C8": "91C8+121D",
                    "1695 9215 2697": "1695+9215+2697", "9215 1695 2697": "1695+9215+2697",
                    "2697 1695 9215": "1695+9215+2697", "2697 9215 1695": "1695+9215+2697"
                },
                "feishu_url": "https://rcni71ddyxpv.feishu.cn/wiki/PPXhwHGiiih4JKkqxwMcYq3PnXe",
                "headers": ["发布date", "Handle（账号名）", "产品名称", "VIDEO LINK（视频链接）", "ADS code（投流码）"]
            }
        }
        for merchant, rule in default_rules.items():
            if merchant not in self.rules:
                self.rules[merchant] = rule
        self.save_rules()

    def get_merchant_list(self):
        return list(self.rules.keys())

    def get_merchant_rule(self, merchant):
        return self.rules.get(merchant, {})

    def add_merchant(self, merchant):
        if merchant in self.rules:
            return False
        self.rules[merchant] = {
            "product_map": {},
            "feishu_url": "",
            "headers": ["日期", "ID", "产品名称", "链接", "VID", "code"]
        }
        self.save_rules()
        return True

    def delete_merchant(self, merchant):
        if merchant not in self.rules:
            return False
        del self.rules[merchant]
        self.save_rules()
        return True

    def update_merchant_rule(self, merchant, product_map=None, feishu_url=None, headers=None):
        if merchant not in self.rules:
            return False
        if product_map is not None:
            self.rules[merchant]["product_map"] = product_map
        if feishu_url is not None:
            self.rules[merchant]["feishu_url"] = feishu_url
        if headers is not None:
            self.rules[merchant]["headers"] = headers
        self.save_rules()
        return True

    def global_convert_product(self, raw_name):
        """全局产品匹配：遍历所有商家规则，返回(归属商家, 转换后名称)"""
        if raw_name in self.global_product_map:
            return self.global_product_map[raw_name]
        # 无匹配则归属为空，名称保留原值
        return ("未知商家", raw_name)


# ====================== 规则配置窗口 ======================
class RuleConfigWindow(tk.Toplevel):
    def __init__(self, parent, rule_manager):
        super().__init__(parent)
        self.parent = parent
        self.rule_mgr = rule_manager
        self.title("规则配置管理")
        self.geometry("1000x700")
        self.resizable(1, 1)
        self.transient(parent)
        self.grab_set()
        try:
            self.iconbitmap(APP_ICON)
        except:
            pass

        main_frame = ttk.Frame(self)
        main_frame.pack(fill="both", expand=1, padx=10, pady=10)
        left_frame = ttk.Frame(main_frame, width=200)
        left_frame.grid(row=0, column=0, sticky="ns", padx=5)
        ttk.Label(left_frame, text="商家列表", font=("微软雅黑", 12, "bold")).pack(pady=5)
        self.merchant_listbox = tk.Listbox(left_frame, width=25, height=30)
        self.merchant_listbox.pack(fill="both", expand=1)
        self.merchant_listbox.bind("<<ListboxSelect>>", self.on_merchant_select)
        btn_frame = ttk.Frame(left_frame)
        btn_frame.pack(fill="x", pady=5)
        ttk.Button(btn_frame, text="新增商家", command=self.add_merchant).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="删除商家", command=self.delete_merchant).pack(side="left", padx=2)

        right_frame = ttk.Frame(main_frame)
        right_frame.grid(row=0, column=1, sticky="nsew", padx=5)
        main_frame.grid_columnconfigure(1, weight=1)
        main_frame.grid_rowconfigure(0, weight=1)
        self.merchant_name_var = tk.StringVar()
        ttk.Label(right_frame, text="商家名称：", font=("微软雅黑", 12)).grid(row=0, column=0, sticky="w", pady=5)
        ttk.Label(right_frame, textvariable=self.merchant_name_var, font=("微软雅黑", 12, "bold")).grid(row=0, column=1, sticky="w", pady=5)
        ttk.Label(right_frame, text="飞书表格链接：").grid(row=1, column=0, sticky="w", pady=5)
        self.feishu_url_entry = ttk.Entry(right_frame, width=80)
        self.feishu_url_entry.grid(row=1, column=1, sticky="ew", pady=5)
        right_frame.grid_columnconfigure(1, weight=1)
        ttk.Label(right_frame, text="目标表头（逗号分隔）：").grid(row=2, column=0, sticky="w", pady=5)
        self.headers_entry = ttk.Entry(right_frame, width=80)
        self.headers_entry.grid(row=2, column=1, sticky="ew", pady=5)
        ttk.Label(right_frame, text="产品名称替换规则", font=("微软雅黑", 12, "bold")).grid(row=3, column=0, columnspan=2, sticky="w", pady=10)
        rule_table_frame = ttk.Frame(right_frame)
        rule_table_frame.grid(row=4, column=0, columnspan=2, sticky="nsew", pady=5)
        right_frame.grid_rowconfigure(4, weight=1)
        self.rule_tree = ttk.Treeview(rule_table_frame, columns=["input", "output"], show="headings", height=20)
        self.rule_tree.heading("input", text="输入的产品名")
        self.rule_tree.heading("output", text="替换后的产品名")
        self.rule_tree.column("input", width=200)
        self.rule_tree.column("output", width=200)
        self.rule_tree.pack(side="left", fill="both", expand=1)
        scrollbar = ttk.Scrollbar(rule_table_frame, orient="vertical", command=self.rule_tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.rule_tree.configure(yscrollcommand=scrollbar.set)
        rule_btn_frame = ttk.Frame(right_frame)
        rule_btn_frame.grid(row=5, column=0, columnspan=2, sticky="w", pady=5)
        ttk.Button(rule_btn_frame, text="新增规则", command=self.add_rule).pack(side="left", padx=2)
        ttk.Button(rule_btn_frame, text="修改规则", command=self.edit_rule).pack(side="left", padx=2)
        ttk.Button(rule_btn_frame, text="删除规则", command=self.delete_rule).pack(side="left", padx=2)
        save_btn_frame = ttk.Frame(right_frame)
        save_btn_frame.grid(row=6, column=0, columnspan=2, sticky="e", pady=10)
        ttk.Button(save_btn_frame, text="保存当前商家规则", command=self.save_merchant_rule, width=20).pack(side="right")
        self.refresh_merchant_list()
        self.current_merchant = None

    def refresh_merchant_list(self):
        self.merchant_listbox.delete(0, tk.END)
        for merchant in self.rule_mgr.get_merchant_list():
            self.merchant_listbox.insert(tk.END, merchant)

    def on_merchant_select(self, event):
        selection = self.merchant_listbox.curselection()
        if not selection:
            return
        merchant = self.merchant_listbox.get(selection[0])
        self.current_merchant = merchant
        self.merchant_name_var.set(merchant)
        rule = self.rule_mgr.get_merchant_rule(merchant)
        self.feishu_url_entry.delete(0, tk.END)
        self.feishu_url_entry.insert(0, rule.get("feishu_url", ""))
        self.headers_entry.delete(0, tk.END)
        self.headers_entry.insert(0, ",".join(rule.get("headers", [])))
        self.rule_tree.delete(*self.rule_tree.get_children())
        for k, v in rule.get("product_map", {}).items():
            self.rule_tree.insert("", "end", values=[k, v])

    def add_merchant(self):
        merchant = simpledialog.askstring("新增商家", "请输入商家名称：", parent=self)
        if not merchant:
            return
        if self.rule_mgr.add_merchant(merchant):
            self.refresh_merchant_list()
            messagebox.showinfo("成功", f"商家【{merchant}】新增成功")
        else:
            messagebox.showerror("错误", "商家已存在")

    def delete_merchant(self):
        selection = self.merchant_listbox.curselection()
        if not selection:
            messagebox.showwarning("提示", "请先选择要删除的商家")
            return
        merchant = self.merchant_listbox.get(selection[0])
        if messagebox.askyesno("确认", f"确定删除【{merchant}】？不可恢复"):
            if self.rule_mgr.delete_merchant(merchant):
                self.refresh_merchant_list()
                self.current_merchant = None
                self.merchant_name_var.set("")
                self.feishu_url_entry.delete(0, tk.END)
                self.headers_entry.delete(0, tk.END)
                self.rule_tree.delete(*self.rule_tree.get_children())
                messagebox.showinfo("成功", "删除成功")

    def add_rule(self):
        if not self.current_merchant:
            messagebox.showwarning("提示", "请先选择商家")
            return
        i = simpledialog.askstring("新增", "输入产品名：", parent=self)
        o = simpledialog.askstring("新增", "替换为：", parent=self)
        if i and o:
            self.rule_tree.insert("", "end", values=[i, o])

    def edit_rule(self):
        s = self.rule_tree.selection()
        if not s:
            messagebox.showwarning("提示", "请选择规则")
            return
        item = self.rule_tree.item(s[0])
        ni = simpledialog.askstring("修改", "产品名：", initialvalue=item["values"][0], parent=self)
        no = simpledialog.askstring("修改", "替换为：", initialvalue=item["values"][1], parent=self)
        if ni and no:
            self.rule_tree.item(s[0], values=[ni, no])

    def delete_rule(self):
        s = self.rule_tree.selection()
        if s and messagebox.askyesno("确认", "删除？"):
            self.rule_tree.delete(s)

    def save_merchant_rule(self):
        if not self.current_merchant:
            messagebox.showwarning("提示", "请先选择商家")
            return
        u = self.feishu_url_entry.get().strip()
        h = [x.strip() for x in self.headers_entry.get().split(",") if x.strip()]
        m = {}
        for i in self.rule_tree.get_children():
            v = self.rule_tree.item(i)["values"]
            if len(v)>=2:
                m[v[0]] = v[1]
        self.rule_mgr.update_merchant_rule(self.current_merchant, m, u, h)
        messagebox.showinfo("成功", "保存成功")


# ====================== 飞书 & 工具函数 ======================
def extract_feishu_excel_id(url):
    match = re.search(r'feishu\.cn/wiki/(\w+)', url)
    return match.group(1) if match else None

def write_to_feishu_table(excel_id, data, headers):
    if not excel_id:
        return False, "链接无效"
    if not data:
        return False, "无数据"
    return True, f"数据准备完成，共{len(data)}行"

def get_beijing_date():
    return datetime.now().strftime("%Y/%m/%d")

def extract_vid_from_link(link):
    match = VID_PATTERN.search(link)
    return match.group(1) if match else ""

# ====================== 网络 ======================
REGION_MAP = {"CN":"大陆","HK":"香港","TW":"台湾","MO":"澳门"}
FORBID = {"CN","HK"}

def get_current_ip():
    try:
        import httpx
        return httpx.get("https://api.ipify.org", timeout=3).text.strip()
    except:
        return None

def check_region():
    try:
        import httpx
        data = httpx.get("https://ipinfo.io/json", timeout=3).json()
        c = data.get("country")
        return data.get("ip"), REGION_MAP.get(c,"海外"), c in FORBID
    except:
        return None, "未知", False

def parse_single_short_url(url):
    if url in url_cache:
        return url_cache[url]
    try:
        import httpx
        global client
        if client is None:
            client = httpx.Client(
                headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"},
                timeout=5,
                follow_redirects=True
            )
        resp = client.get(url, follow_redirects=True)
        final_url = str(resp.url).split("?")[0]
        url_cache[url] = final_url
        return final_url
    except Exception as e:
        print(f"解析失败: {url}, 错误: {e}")
        url_cache[url] = url
        return url


# ====================== 解析核心（全局产品匹配） ======================
def background_parse_task(text, stop_flag, callback, rule_mgr):
    import httpx
    import ssl
    from concurrent.futures import ThreadPoolExecutor, as_completed
    global client

    ssl._create_default_https_context = ssl._create_unverified_context
    if client is None:
        client = httpx.Client(
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"},
            timeout=5,
            follow_redirects=True
        )

    start_time = time.perf_counter()
    raw_lines = [line.strip() for line in text.splitlines() if line.strip()]
    tasks = []
    index = 0
    current_account = None
    total_lines = len(raw_lines)

    while index < total_lines:
        if stop_flag.is_set():
            cost = round(time.perf_counter() - start_time, 2)
            callback([], f"⏹️ 解析停止 | 耗时 {cost}s")
            return
        line = raw_lines[index]
        if index + 3 <= total_lines and TIKTOK_SHORT_PATTERN.match(raw_lines[index + 2]) and ADS_CODE_PATTERN.match(raw_lines[index + 3]):
            current_account = line
            index += 1
            while index + 2 <= total_lines:
                if stop_flag.is_set():
                    cost = round(time.perf_counter() - start_time, 2)
                    callback([], f"⏹️ 解析停止 | 耗时 {cost}s")
                    return
                prod = raw_lines[index]
                url = raw_lines[index + 1]
                code = raw_lines[index + 2]
                if TIKTOK_SHORT_PATTERN.match(url) and ADS_CODE_PATTERN.match(code):
                    tasks.append((current_account, prod, url, code))
                    index += 3
                else:
                    break
            continue
        index += 1

    if not tasks:
        cost = round(time.perf_counter() - start_time, 2)
        callback([], f"❌ 无有效数据 | 耗时 {cost}s")
        return

    result = []
    parse_date = get_beijing_date()
    max_workers = min(16, len(tasks))

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_task = {executor.submit(parse_single_short_url, t[2]): t for t in tasks}
        for future in as_completed(future_to_task):
            if stop_flag.is_set():
                cost = round(time.perf_counter() - start_time, 2)
                callback([], f"⏹️ 解析停止 | 耗时 {cost}s")
                return
            account, raw_product, short_url, ad_code = future_to_task[future]
            final_url = future.result()
            vid = extract_vid_from_link(final_url)

            # 核心1：全局所有商家规则匹配产品，获取归属商家+转换后名称
            belong_merchant, new_product = rule_mgr.global_convert_product(raw_product)

            # 数据结构：[日期, 账号ID, 转换后产品, 链接, VID, code, 原始产品, 归属商家]
            result.append([
                parse_date, account, new_product, final_url, vid, ad_code, raw_product, belong_merchant
            ])

    cost = round(time.perf_counter() - start_time, 2)
    callback(result, f"✅ 解析完成 {len(result)} 条 | 耗时 {cost} 秒")


# ====================== 主界面 ======================
class TikTokToolGUI:
    def __init__(self, root, auth_mgr, rule_mgr):
        self.root = root
        self.auth_mgr = auth_mgr
        self.rule_mgr = rule_mgr
        self.root.title(APP_TITLE)
        self.root.geometry("1000x750")
        self.full_data = []       # 完整原始数据（含归属商家）
        self.stop_flag = threading.Event()
        self.is_parsing = False
        self.net_forbidden = False
        self.selected_merchant = None
        self.parse_result_msg = ""
        self.last_refresh = 0
        self.forbid_popup = False

        main = ttk.Frame(root)
        main.pack(fill="both", expand=1, padx=10, pady=10)

        # 网络行
        netf = ttk.Frame(main)
        netf.grid(row=0, column=0, sticky="ew", pady=2)
        self.netv = tk.StringVar(value="检测中…")
        ttk.Label(netf, text="网络：").pack(side="left")
        ttk.Label(netf, textvariable=self.netv, foreground="blue").pack(side="left")
        ttk.Button(netf, text="刷新网络", command=self.ref_net).pack(side="right")

        # 商家选择行
        merf = ttk.Frame(main)
        merf.grid(row=1, column=0, sticky="w", pady=2)
        ttk.Label(merf, text="当前选中商家：").pack(side="left", padx=5)
        self.mer = ttk.Combobox(merf, state="readonly", width=20)
        self.mer.pack(side="left", padx=5)
        self.mer.bind("<<ComboboxSelected>>", self.on_mer)
        ttk.Button(merf, text="规则配置", command=self.open_rule).pack(side="left", padx=5)

        # 自动导出飞书
        auto_frame = ttk.Frame(main)
        auto_frame.grid(row=1, column=0, sticky="e", pady=2)
        self.auto = tk.BooleanVar(value=True)
        ttk.Checkbutton(auto_frame, text="自动导出飞书", variable=self.auto).pack(side="right")

        # 输入区提示
        ttk.Label(main, text="格式：ID → 产品名称+链接+code(可连续多组)，支持任意空行").grid(row=2, column=0, sticky="w")
        self.txt = tk.Text(main, wrap="word")
        self.txt.grid(row=3, column=0, sticky="nsew", pady=5)

        # 功能按钮
        btf = ttk.Frame(main)
        btf.grid(row=4,column=0,sticky="w")
        self.br = ttk.Button(btf,text="一键解析",command=self.start)
        self.br.pack(side="left",padx=5)
        self.bs = ttk.Button(btf,text="停止",command=self.stop,state="disabled")
        self.bs.pack(side="left",padx=5)
        ttk.Button(btf,text="导出Excel",command=self.export_excel).pack(side="left",padx=5)
        ttk.Button(btf,text="导出飞书",command=self.export_feishu).pack(side="left",padx=5)
        ttk.Button(btf,text="清空",command=self.clear_all).pack(side="left",padx=5)

        # 预览标题
        ttk.Label(main,text=f"数据预览 | {auth_mgr.get_remain_days()}").grid(row=5,column=0,sticky="w")
        # 预览表格固定基础列
        cols = ["date","id","prod","link","vid","code"]
        self.tree = ttk.Treeview(main,columns=cols,show="headings")
        self.tree.heading("date",text="日期")
        self.tree.heading("id",text="账号ID")
        self.tree.heading("prod",text="转换后产品")
        self.tree.heading("link",text="视频链接")
        self.tree.heading("vid",text="VID")
        self.tree.heading("code",text="投流CODE")
        self.tree.column("date",width=80)
        self.tree.column("id",width=110)
        self.tree.column("prod",width=140)
        self.tree.column("link",width=320)
        self.tree.column("vid",width=160)
        self.tree.column("code",width=160)
        self.tree.grid(row=6,column=0,sticky="nsew",pady=5)

        # 状态栏
        stf = ttk.Frame(main)
        stf.grid(row=7,column=0,sticky="ew")
        self.stat = tk.StringVar(value="就绪")
        ttk.Label(stf,textvariable=self.stat,foreground="green").pack(side="left")

        # 权重分配
        main.grid_rowconfigure(3,weight=2)
        main.grid_rowconfigure(6,weight=5)
        main.grid_columnconfigure(0,weight=1)

        self.ref_mer()
        self.root.after(200, self.ref_net)
        self.root.bind("<FocusIn>", self.on_focus)

    def ref_mer(self):
        m = self.rule_mgr.get_merchant_list()
        self.mer["values"] = m
        if m:
            self.mer.current(0)
            self.selected_merchant = m[0]

    def on_mer(self,e):
        self.selected_merchant = self.mer.get()

    def open_rule(self):
        RuleConfigWindow(self.root, self.rule_mgr)
        self.ref_mer()

    def on_focus(self,e=None):
        if self.is_parsing: return
        if time.time()-self.last_refresh < 4: return
        self.last_refresh = time.time()
        self.ref_net()

    def ref_net(self):
        self.stat.set("检测网络…")
        def t():
            ip,reg,forb = check_region()
            self.net_forbidden = forb
            self.root.after(0,lambda: self.netv.set(f"{ip} | {reg}"))
            self.root.after(0,lambda: self.stat.set("就绪"))
            self.root.after(0,lambda: self.br.config(state="disabled" if forb else "normal"))
            if forb and not self.forbid_popup:
                self.forbid_popup = True
                self.root.after(0,lambda: messagebox.showwarning("禁止","大陆/香港网络不可使用"))
        threading.Thread(target=t,daemon=True).start()

    def start(self):
        if self.is_parsing or self.net_forbidden:
            return
        s = self.txt.get("1.0","end-1c").strip()
        if not s:
            messagebox.showwarning("提示","请输入解析内容")
            return
        self.is_parsing = True
        self.stop_flag.clear()
        self.br.config(state="disabled")
        self.bs.config(state="normal")
        self.stat.set("解析中…")

        def cb(res,msg):
            self.full_data = res
            # 刷新预览表格
            for item in self.tree.get_children():
                self.tree.delete(item)
            for row in res:
                # 预览只展示基础6列
                self.tree.insert("","end",values=row[:6])
            self.parse_result_msg = msg
            self.stat.set(msg)
            self.br.config(state="normal")
            self.bs.config(state="disabled")
            self.is_parsing = False
            # 自动导出飞书
            if self.auto.get() and res:
                self.export_feishu()
                self.stat.set(self.parse_result_msg)

        # 不再传选中商家，使用全局匹配
        threading.Thread(target=background_parse_task,args=(s,self.stop_flag,cb,self.rule_mgr),daemon=True).start()

    def stop(self):
        if not self.is_parsing:
            return
        self.stop_flag.set()
        self.stat.set("停止中…")

    def export_excel(self):
        if not self.full_data:
            messagebox.showwarning("提示","暂无解析数据")
            return
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件","*.xlsx")]
        )
        if not save_path:
            return

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # 核心2：按归属商家聚合数据
        merchant_group = {}
        for row in self.full_data:
            belong_mer = row[7]
            if belong_mer not in merchant_group:
                merchant_group[belong_mer] = []
            merchant_group[belong_mer].append(row)

        # 字段映射：表头名称 -> 数据下标
        # 修改后的代码
        field_index_map = {
            "日期": 0, "发布date": 0, "发布日期": 0,  # 补上"发布日期"
            "ID": 1, "账号ID": 1, "handle": 1, "Handle（账号名）": 1,
            "产品名称": 2, "PID(or型号)": 2,
            "链接": 3, "VIDEO LINK（视频链接）": 3, "视频链接": 3,  # 补上"视频链接"
            "VID": 4, "vid": 4,
            "code": 5, "ADS code（投流码）": 5, "投流码": 5,
            "挂链接店铺名称": 6
        }

        # 逐个商家生成独立Sheet
        for mer_name, data_list in merchant_group.items():
            ws = wb.create_sheet(title=mer_name)
            rule = self.rule_mgr.get_merchant_rule(mer_name)
            headers = rule.get("headers", ["日期", "ID", "产品名称", "链接", "VID", "code"])

            # 写入表头
            for col_idx, head_name in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=head_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="CCCCCC")

            # 逐行写入数据（按当前商家表头顺序取值）
            for row_idx, data_row in enumerate(data_list, 2):
                for col_idx, head_name in enumerate(headers, 1):
                    idx = field_index_map.get(head_name, -1)
                    val = data_row[idx] if idx != -1 else ""
                    ws.cell(row=row_idx, column=col_idx, value=val)

            # 列宽自适应
            for c in range(1, len(headers)+1):
                ws.column_dimensions[get_column_letter(c)].width = 16

        wb.save(save_path)
        messagebox.showinfo("导出成功", f"文件已保存至：{save_path}")

    def export_feishu(self):
        if not self.full_data:
            messagebox.showwarning("提示","暂无解析数据")
            return
        # 飞书使用当前选中商家配置
        rule = self.rule_mgr.get_merchant_rule(self.selected_merchant)
        url = rule.get("feishu_url")
        heads = rule.get("headers",[])
        if not url or not heads:
            messagebox.showerror("错误","请先配置当前商家的飞书链接与表头")
            return
        eid = extract_feishu_excel_id(url)
        # 修改后的代码
        field_index_map = {
            "日期": 0, "发布date": 0, "发布日期": 0,  # 补上"发布日期"
            "ID": 1, "账号ID": 1, "handle": 1, "Handle（账号名）": 1,
            "产品名称": 2, "PID(or型号)": 2,
            "链接": 3, "VIDEO LINK（视频链接）": 3, "视频链接": 3,  # 补上"视频链接"
            "VID": 4, "vid": 4,
            "code": 5, "ADS code（投流码）": 5, "投流码": 5,
            "挂链接店铺名称": 6
        }
        fmt_data = []
        for row in self.full_data:
            temp = []
            for h in heads:
                idx = field_index_map.get(h, -1)
                temp.append(row[idx] if idx != -1 else "")
            fmt_data.append(temp)

        self.stat.set("正在同步飞书…")
        def t():
            ok,info = write_to_feishu_table(eid, fmt_data, heads)
            self.root.after(0,lambda: self.stat.set(self.parse_result_msg))
            self.root.after(0,lambda: messagebox.showinfo("完成" if ok else "失败", info))
        threading.Thread(target=t,daemon=True).start()

    def clear_all(self):
        self.txt.delete("1.0","end")
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.full_data = []
        self.stat.set("已清空")


# ====================== 主函数 ======================
def main():
    root = tk.Tk()
    root.withdraw()
    am = AuthManager()
    rm = RuleManager()
    try:
        root.iconbitmap(APP_ICON)
    except:
        pass
    def ui():
        if not am.valid:
            ActivateWindow(root,am)
        else:
            TikTokToolGUI(root,am,rm)
        root.deiconify()
    root.after(1, ui)
    def on_close():
        global client
        if client:
            try:
                client.close()
            except:
                pass
        root.destroy()
    root.protocol("WM_DELETE_WINDOW", on_close)
    root.mainloop()

if __name__ == "__main__":
    main()